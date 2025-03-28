from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Max
from openpyxl.styles import PatternFill

from openpyxl.styles.colors import WHITE
import array as arr

@csrf_exempt
def export_tracker(request):
	current_date  = datetime.now().date()
	year          = datetime.strptime(str(current_date), "%Y-%m-%d").strftime('%Y')
	try:
		_date 		= json.loads(request.GET['entry_date'])
		split_date 	= datetime.strptime(_date, "%d/%b/%Y").strftime('%Y-%m-%d')
		entry_date 	= split_date		
	except:
		entry_date 	= ''

	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/weekly_audit.xlsx'
	dstroot       = select_dir+'/Weekly_Audit/weekly_audit.xlsx'	 
	copyfile(srcfile, dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/Weekly_Audit/weekly_audit.xlsx')
	wb 	   	      = load_workbook(my_dir+'/Weekly_Audit/weekly_audit.xlsx')
	ws     	      = wb.get_sheet_by_name("Sheet1")	
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	arrr = []
	arr1 = []
	i    = 3
	j    = 3
	
	curr_date      = datetime.now().date()	
	userID 		   = models.User_entry.objects.aggregate(Max('id'))
	max_userID     = userID['id__max']

	if entry_date:
		ws.cell('B1').value = datetime.strptime(entry_date, "%Y-%m-%d").strftime('%d/%b/%Y')
		ws.cell('E1').value = datetime.strptime(entry_date, "%Y-%m-%d").strftime('%d/%b/%Y')
	else:
		ws.cell('B1').value = ""
		ws.cell('E1').value = ""

	tot_sys		 	= models.weekly_entry.objects.filter(entry_date=entry_date).count()
	split_print 	= tot_sys/2	
	break_print     = int(split_print)

	tot_sys1		 	= models.weekly_entry.objects.filter(entry_date=entry_date).aggregate(Min('id'))
	tot_sys2		 	= models.weekly_entry.objects.filter(entry_date=entry_date).aggregate(Max('id'))
	min_ids 			= tot_sys1['id__min']
	print '_____________>>',tot_sys2['id__max']
	system_details 	= models.weekly_entry.objects.filter(entry_date=entry_date).order_by('user__user_name')

	_styles 		= Style(font=Font(name='Calibri', size=11, bold=True),alignment=Alignment(horizontal='center'),fill=PatternFill(patternType='solid', fgColor=Color(rgb='C0C0C0')),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	_normal 		= Style(font=Font(name='Calibri', size=11, bold=False),alignment=Alignment(horizontal='left'),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	_right_tick     = Style(font=Font(name='Wingdings 2', size=11, bold=True),alignment=Alignment(horizontal='center'),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	_wrong_tick     = Style(font=Font(name='Wingdings 2', size=11, bold=True),alignment=Alignment(horizontal='center'),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	_for_center     = Style(font=Font(name='Wingdings 2', size=11, bold=True),alignment=Alignment(horizontal='center'),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	_for_bck        = Style(font=Font(name='Calibri', size=11, bold=True),alignment=Alignment(horizontal='center'),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	
	for t in system_details:		
		try:			
			ws.cell('A'+str(i)).value 	= t.user.user_name
			
		except:
			ws.cell('A'+str(i)).value 	= "-"

		if t.unused_program=='YES':	
			ws.cell('B'+str(i+2)).value = 'Yes'						
		else:
			ws.cell('B'+str(i+2)).value = '-'			
			
		if t.temp_file=='YES':	
			ws.cell('B'+str(i+3)).value = 'Yes'			
		else:
			ws.cell('B'+str(i+3)).value = '-'			
			
		if t.up_antivirus=='YES':	
			ws.cell('B'+str(i+4)).value = 'Yes'			
		else:
			ws.cell('B'+str(i+4)).value = '-'			
			ws.cell('B'+str(i+4)).value = '-'	
			
		if t.up_antivirus=='YES':	
			ws.cell('B'+str(i+4)).value = 'Yes'			
		else:
			ws.cell('B'+str(i+4)).value = '-'			
			ws.cell('B'+str(i+4)).value = '-'	
			
		if t.history_clean=='YES':	
			ws.cell('B'+str(i+5)).value = 'Yes'		
		else:
			ws.cell('B'+str(i+5)).value = '-'			
			ws.cell('B'+str(i+5)).value = '-'	
				
		if t.server_folder=='YES':	
			ws.cell('B'+str(i+6)).value = 'Yes'			
		else:
			ws.cell('B'+str(i+6)).value = '-'			
			ws.cell('B'+str(i+6)).value = '-'	
				
		if t.internet_conn=='YES':	
			ws.cell('B'+str(i+7)).value = 'Yes'			
		else:
			ws.cell('B'+str(i+7)).value = '-'			
			ws.cell('B'+str(i+7)).value = '-'	
				
		if t.printer_conn=='YES':	
			ws.cell('B'+str(i+8)).value = 'Yes'			
		else:
			ws.cell('B'+str(i+8)).value = '-'	
			ws.cell('B'+str(i+8)).value = '-'
			
		ws.cell('B'+str(i+9)).value = t.email_configure #t.issue_any
		ws.cell('B'+str(i+9)).style = _normal	


		ws.cell('A'+str(i+1)).value = 'Boot Time'
		ws.cell('A'+str(i+2)).value = 'Unused program removed?'
		ws.cell('A'+str(i+3)).value = 'Temp files removed? Includes .tmp & .chk files'
		ws.cell('A'+str(i+4)).value = 'Latest anti-virus update 1 day old?'
		ws.cell('A'+str(i+5)).value = 'History & Cache file cleaned up?'
		ws.cell('A'+str(i+6)).value = 'NAS/Server folder connectivity (Yes/No)'
		ws.cell('A'+str(i+7)).value = 'Internet Connectivity (Yes/No)'
		ws.cell('A'+str(i+8)).value = 'Printer Connectivity (Yes/No)'
		ws.cell('A'+str(i+9)).value = 'Issue, if any (User Excel comment)'
		ws.cell('B'+str(i+1)).value = t.boot_time
		ws.cell('B'+str(i+1)).style = _normal
		ws.cell('A'+str(i)).style 	= _styles	
		ws.cell('B'+str(i)).style 	= _styles			
		ws.cell('B'+str(i+1)).style = _for_bck	
		ws.cell('A'+str(i+2)).style = _normal	
		ws.cell('A'+str(i+3)).style = _normal	
		ws.cell('A'+str(i+4)).style = _normal	
		ws.cell('A'+str(i+5)).style = _normal	
		ws.cell('A'+str(i+6)).style = _normal	
		ws.cell('A'+str(i+7)).style = _normal	
		ws.cell('A'+str(i+8)).style = _normal	
		ws.cell('A'+str(i+9)).style = _normal			
		i+=10
		
	# 	if t.id not in arrr:
	# 		arrr.append(t.id)
	# 	ids = max(arrr)
	# arr = (arrr)
	

	# for v in arr:	
		
	# 	system_deta    = models.weekly_entry.objects.filter(entry_date=entry_date,id__gt=v).order_by('user__user_name')[:int(break_print)]
	# 	_styles 		= Style(font=Font(name='Calibri', size=11, bold=True),alignment=Alignment(horizontal='center'),fill=PatternFill(patternType='solid', fgColor=Color(rgb='C0C0C0')),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	# 	_normal 		= Style(font=Font(name='Calibri', size=11, bold=False),alignment=Alignment(horizontal='left'),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	# 	_for_bckx       = Style(font=Font(name='Calibri', size=11, bold=True),alignment=Alignment(horizontal='center'),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
		
	# 	k = 3
	# 	l = 3
	# 	for tt in system_deta:
	# 		try:			
	# 			ws.cell('D'+str(k)).value 	= tt.user.user_name				
	# 		except:
	# 			ws.cell('D'+str(k)).value 	= "-"

	# 		if tt.unused_program=='YES':	
	# 			ws.cell('E'+str(k+2)).value = 'Yes'						
	# 		else:
	# 			ws.cell('E'+str(k+2)).value = '-'			
				
	# 		if tt.temp_file=='YES':	
	# 			ws.cell('E'+str(k+3)).value = 'Yes'			
	# 		else:
	# 			ws.cell('E'+str(k+3)).value = '-'			
				
	# 		if tt.up_antivirus=='YES':	
	# 			ws.cell('E'+str(k+4)).value = 'Yes'			
	# 		else:
	# 			ws.cell('E'+str(k+4)).value = '-'			
	# 			ws.cell('E'+str(k+4)).value = '-'	
				
	# 		if tt.up_antivirus=='YES':	
	# 			ws.cell('E'+str(k+4)).value = 'Yes'			
	# 		else:
	# 			ws.cell('E'+str(k+4)).value = '-'			
	# 			ws.cell('E'+str(k+4)).value = '-'	
				
	# 		if tt.history_clean=='YES':	
	# 			ws.cell('E'+str(k+5)).value = 'Yes'		
	# 		else:
	# 			ws.cell('E'+str(k+5)).value = '-'			
	# 			ws.cell('E'+str(k+5)).value = '-'	
					
	# 		if tt.server_folder=='YES':	
	# 			ws.cell('E'+str(k+6)).value = 'Yes'			
	# 		else:
	# 			ws.cell('E'+str(k+6)).value = '-'			
	# 			ws.cell('E'+str(k+6)).value = '-'	
					
	# 		if tt.internet_conn=='YES':	
	# 			ws.cell('E'+str(k+7)).value = 'Yes'			
	# 		else:
	# 			ws.cell('E'+str(k+7)).value = '-'			
	# 			ws.cell('E'+str(k+7)).value = '-'	
					
	# 		if tt.printer_conn=='YES':	
	# 			ws.cell('E'+str(k+8)).value = 'Yes'			
	# 		else:
	# 			ws.cell('E'+str(k+8)).value = '-'	
	# 			ws.cell('E'+str(k+8)).value = '-'
				
	# 		ws.cell('E'+str(k+9)).value = tt.issue_any		
	# 		ws.cell('D'+str(k+1)).value = 'Boot Time'
	# 		ws.cell('D'+str(k+2)).value = 'Unused program removed?'
	# 		ws.cell('D'+str(k+3)).value = 'Temp files removed? Includes .tmp & .chk files'
	# 		ws.cell('D'+str(k+4)).value = 'Latest anti-virus update 1 day old?'
	# 		ws.cell('D'+str(k+5)).value = 'History & Cache file cleaned up?'
	# 		ws.cell('D'+str(k+6)).value = 'NAS/Server folder connectivity (Yes/No)'
	# 		ws.cell('D'+str(k+7)).value = 'Internet Connectivity (Yes/No)'
	# 		ws.cell('D'+str(k+8)).value = 'Printer Connectivity (Yes/No)'
	# 		ws.cell('D'+str(k+9)).value = 'Issue, if any (User Excel comment)'
	# 		ws.cell('E'+str(k+1)).value = tt.boot_time	
	# 		ws.cell('E'+str(j+1)).style = _for_bckx
	# 		ws.cell('D'+str(j)).style   = _styles	
	# 		ws.cell('E'+str(j)).style 	= _styles		
	# 		ws.cell('D'+str(j+1)).style = _normal	
	# 		ws.cell('D'+str(j+2)).style = _normal	
	# 		ws.cell('D'+str(j+3)).style = _normal	
	# 		ws.cell('D'+str(j+4)).style = _normal	
	# 		ws.cell('D'+str(j+5)).style = _normal	
	# 		ws.cell('D'+str(j+6)).style = _normal	
	# 		ws.cell('D'+str(j+7)).style = _normal	
	# 		ws.cell('D'+str(j+8)).style = _normal	
	# 		ws.cell('D'+str(j+9)).style = _normal		
	# 		k+=10			

	
		wb.save(my_dir+'/Weekly_Audit/weekly_audit.xlsx')
	return HttpResponse(json.dumps('done'))

def weekly_audit_export(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/Weekly_Audit/weekly_audit.xlsx')
	filename 	   					= my_dir+'/Weekly_Audit/weekly_audit.xlsx'	
	download_name  					= "weekly_audit.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response

def export_password_share(request):
	current_date  = datetime.now().date()
	year          = datetime.strptime(str(current_date), "%Y-%m-%d").strftime('%Y')
	try:
		_date 		= json.loads(request.GET['entry_date'])
		split_date 	= datetime.strptime(_date, "%d/%b/%Y").strftime('%Y-%m-%d')
		entry_date 	= split_date		
	except:
		entry_date 	= ''

	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/Password_list.xlsx'
	dstroot       = select_dir+'/Password_list.xlsx'	 
	copyfile(srcfile, dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/Password_list.xlsx')
	wb 	   	      = load_workbook(my_dir+'/Password_list.xlsx')
	ws     	      = wb.get_sheet_by_name("Sheet1")	
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	arrr = []
	arr1 = []
	i    = 2
	j    = 2
	t=0
	sys_details   = models.weekly_entry.objects.filter(entry_date=current_date)
	for x in sys_details:		
		ws.cell('A'+str(i)).value 	= (t+1)
		ws.cell('B'+str(i)).value 	= x.user.user_name
		ws.cell('C'+str(i)).value 	= x.issue_any
		i+=1
		j+=1
		t+=1
		wb.save(my_dir+'/Password_list.xlsx')
	return HttpResponse(json.dumps('done'))

def download_password_share(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/Password_list.xlsx')
	filename 	   					= my_dir+'/Password_list.xlsx'	
	download_name  					= "Password List.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response


def load_selected_date(request):
	audit_details = models.weekly_entry.objects.all().order_by('-entry_date')
	audit_array   = []
	for s in audit_details:
		if s.entry_date not in audit_array:
			audit_array.append(s.entry_date)

	date_filter = []
	for c in audit_array:
		try:
			df = datetime.strptime(str(c),"%Y-%m-%d").strftime('%d/%b/%Y')
		except:
			df = ''
		date_filter.append({
			'date_filter' : df
			})	
	
	context={
	'cal_details' : date_filter
	}			
	#print '========',date_filter
	return HttpResponse(json.dumps(context))
		

def system_export(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/system_details.xlsx')
	filename 	   					= my_dir+'/system_details.xlsx'	
	download_name  					= "system_details.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response


@csrf_exempt
def export_system_details(request):
	from shutil import copyfile
	import shutil

	fleet_details = json.loads(request.POST['user_details'])
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/system_details.xlsx'
	dstroot       = select_dir+'/system_details.xlsx'
	copyfile(srcfile, dstroot)
	my_dir 	      = os.path.dirname(__file__)
	location      = (my_dir+'/system_details.xlsx')
	wb 	   	      = load_workbook(my_dir+'/system_details.xlsx')
	ws     	      = wb.get_sheet_by_name("sys")
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Arial', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='right'))
	align_style   = Style(font=Font(name='Arial', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='left'))	 	
	file_name     = 'system_details.xlsx'
	
	i = 3
	j = 1
	for ft in fleet_details:			
		system_name   		  	  = ft['System Name']
		processors 	  			  = ft['Processor']		
		ram_details  			  = ft['RAM']	
		hdd_details 		  	  = ft['HDD']
		keyboard_details 		  = ft['KEYBOARD']
		mouse_details 		      = ft['MOUSE']
		monitor_details 		  = ft['MONITOR']
		ups_details 			  = ft['UPS']
		operating_system 		  = ft['Operating System']
		computer_types 			  = ft['Computer Types']
		remarks_details 		  = ft['Remarks']
		types 		  			  = ft['Computer Types']
		locations 				  = ft['Location']
		assign_to 				  = ft['Assigned to ?']
		#department_details 		  = models.weekly_entry.objects.filter(user__user_name=system_name).first()

		ws.cell('A'+str(i)).value = j
		ws.cell('B'+str(i)).value = system_name
		try:
			ws.cell('C'+str(i)).value = locations #department_details.depart_name
		except:
			ws.cell('C'+str(i)).value = ''

		ws.cell('D'+str(i)).value = hdd_details		
		ws.cell('E'+str(i)).value = processors
		ws.cell('F'+str(i)).value = ram_details
		ws.cell('G'+str(i)).value = operating_system
		ws.cell('H'+str(i)).value = remarks_details
		if keyboard_details!='None':
			ws.cell('I'+str(i)).value = keyboard_details
		else:
			ws.cell('I'+str(i)).value = ""
		
		if monitor_details!='None':
			ws.cell('J'+str(i)).value = monitor_details
		else:
			ws.cell('J'+str(i)).value = ""
		
		if mouse_details!='None':
			ws.cell('K'+str(i)).value = mouse_details
		else:
			ws.cell('K'+str(i)).value = ""

		if ups_details!='None':
			ws.cell('L'+str(i)).value = ups_details
		else:
			ws.cell('L'+str(i)).value = ""

		if types!='None':
			ws.cell('M'+str(i)).value = types
		else:
			ws.cell('M'+str(i)).value = ""

		if assign_to!='None':
			ws.cell('N'+str(i)).value = assign_to
		else:
			ws.cell('N'+str(i)).value = ""

		ws.cell('A'+str(i)).style = align_style
		ws.cell('B'+str(i)).style = align_style
		ws.cell('C'+str(i)).style = align_style
		ws.cell('D'+str(i)).style = align_style
		ws.cell('E'+str(i)).style = align_style
		ws.cell('F'+str(i)).style = align_style
		ws.cell('G'+str(i)).style = align_style	
		ws.cell('H'+str(i)).style = align_style	
		ws.cell('I'+str(i)).style = align_style		
		ws.cell('J'+str(i)).style = align_style
		ws.cell('K'+str(i)).style = align_style
		ws.cell('L'+str(i)).style = align_style
		ws.cell('M'+str(i)).style = align_style
		ws.cell('N'+str(i)).style = align_style
		i+=1
		j+=1
	j+=1

	wb.save(my_dir+'/'+file_name)
	return HttpResponse(json.dumps('done'))


def show_image(request):
	
	return HttpResponse(json.dumps('done'))