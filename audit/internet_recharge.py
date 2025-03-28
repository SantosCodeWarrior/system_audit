from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import WHITE


def internet_recharge(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		try:
			recharge_details 	= models.notification_details.objects.all()
			rech_arr 			= []
			rech_today 			= datetime.now().date()
			plan_type 			= []
			alert_Array 		= []			
			for c in recharge_details:
				alert_statuz = (c.alert_date-rech_today).days
				follow_up    = str(c.alert_date)
				plan_namex   = c.plans.plan_type				
				if alert_statuz<=2:
					#print '------',alert_statuz,'---',plan_namex
					log 		  	= request.GET['log']								
					text_content   	= 'This is an important message.'	
					html_content   	= 'Good day Sir,<br><br>This is just a polite reminder.<br><font size="3em"><br><br>'+(log)+'</font><br>From:<br><br>System Auditor'	
					mailto 		   	= 'rakesh.chauhan@bwesglobal.com' #capt.kumaresh@bwesglobal.com
					cc_email 	   	= 'santosh@bwesglobal.com'					
					bcc_email 		= 'system.auditor@bwesglobal.com'
					mailsubject    	= 'ALERT: Polite Reminder for recharge'
					msg  		   	= EmailMultiAlternatives(mailsubject,text_content,'system.auditor@bwesglobal.com',[mailto],bcc=[bcc_email],cc=[cc_email])		
					msg.attach_alternative(html_content, "text/html")	
					msg.send()	
				print '-----Sent 2'
				
		except:
			log = ''

		recharge_dt 	 = models.notification_details.objects.all()
		for w in recharge_dt:		
			curr_today 	 = datetime.now().date()
			alert_statux = (w.alert_date-curr_today).days		
			#print '------',alert_statux
			if alert_statux<=2:
				payment_status 	= w.payment_status
				amt 			= w.amount 
				alert_date 		= w.alert_date.strftime('%d-%b-%Y')
				remaining_day 	= alert_statux
				plan_name    	= w.plans.plan_type
				if alert_statux==1:
					status 			= 'Remained 1 day'
					color_code 		= '#fa9696'
					acive_color 	= 'red'
				elif alert_statux<1:
					status 			= 'Recharge'
					color_code 		= '#fa9696'
					acive_color 	= 'red'
				else:
					status 			= 'Active'
					color_code 		= 'white'
					acive_color 	= 'green'

				try:
					s = c.status
				except:
					s = ''





				alert_Array.append({
					'plan_name' 		: plan_name,
					'payment_status' 	: payment_status,
					'amount'			: amt,
					'alert_date'		: str(alert_date),
					'remaining_day'     : remaining_day,
					'remarks'			: w.remarks,
					'status'			: status,
					'id' 				: w.id,
					'check_status'      : s,
					})
		
		
		plan_details 	= models.notification_details.objects.all()
		plan_arr 		= []
		today 			= datetime.now().date()
		
		for c in plan_details:
			alert_status = (c.alert_date-today).days			
			follow_up    = str(c.alert_date)
			plan_name    = c.plans.plan_type
			alert_date   = c.alert_date
			
			if alert_status==1:
				status 			= 'Remained 1 day'
				color_code 		= '#fa9696'
				acive_color 	= 'red'
			elif alert_status<=2:
				status 			= 'Recharge'
				color_code 		= '#fa9696'
				acive_color 	= 'red'
			else:
				status 			= 'Active'
				color_code 		= 'white'
				acive_color 	= 'green'

			try:
				upload_doc = '/static/PDF/'+str(c.doc_link)#'/static/PDF/UK07BY2882.pdf'
				
			except:
				upload_doc = ''
				

			if c.doc_link!=None:
				doc_label = 'View'
			else:
				doc_label = ''

			print '---------',c.doc_link

			plan_arr.append({
				'follow_up'    	: (c.alert_date).strftime('%d-%b-%Y'),
				'plan_name'    	: c.plans.plan_type,
				'alert_date'   	: c.alert_date,
				'alert_status' 	: str(alert_status),
				'today' 		: today.strftime('%d-%b-%Y'),
				'pay_status'    : c.payment_status,
				'status' 		: status,
				'color_code'    : color_code,
				'active_color'  : acive_color,
				'amount'		: c.amount,
				'id'			: c.id,
				'remind_date'	: str(c.alert_date),
				'remarks'		: c.remarks,
				'statuse'		: c.status,
				'doc_label' 	: doc_label,
				'upload_doc'    : upload_doc,		
				})

		all_plan = models.plan_details.objects.all()
		for t in all_plan:
			plan_type.append({
				'id' 		: t.id,
				'plan_name' : t.plan_type
				})

		context={
		'plan_arr'  	: plan_arr,
		'plan_type' 	: plan_type,
		'alert_Array' 	: alert_Array
		}
		return render_to_response("audit_display/internet_recharge.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')



def update_remind_entry(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		selected_id 		= json.loads(request.GET['selected_id'])
		get_remind_date 	= json.loads(request.GET['get_remind_date'])
		ex_amount 			= json.loads(request.GET['ex_amount'])
		ex_remarks 			= json.loads(request.GET['ex_remarks'])
		ex_statuss 			= json.loads(request.GET['ex_statuss'])
		
		db 					= models.notification_details.objects.filter(id=selected_id).first()
		db.alert_date 		= str(get_remind_date)
		db.amount 			= ex_amount
		db.remarks 			= ex_remarks
		db.status 			= ex_statuss		
		db.save()

		return HttpResponse(json.dumps('done'))
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


def insert_remind_entry(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		
		ee_recharge 	 = json.loads(request.GET['ee_recharge'])		
		get_plan_details = json.loads(request.GET['get_plan_details'])
		ee_items 		 = json.loads(request.GET['ee_items'])
		ee_amount 		 = json.loads(request.GET['ee_amount'])
		

		check =  models.notification_details.objects.filter(plans_id=get_plan_details).count()
		
		if check>1:
			return HttpResponse(json.dumps('plz'))
		else:
			db =  models.notification_details()

		db.alert_date 		= ee_recharge
		db.payment_status 	= get_plan_details
		db.amount 			= ee_amount
		db.plans_id 		= int(get_plan_details)
		db.sent_date 		= datetime.now().date()
		db.email_name		= 'santosh@bwesglobal.com'
		db.save()

		return HttpResponse(json.dumps('done'))
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


def add_plan_entry(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		ee_plan_type = json.loads(request.GET['ee_plan_type'])
		CHK 		 = models.plan_details.objects.filter(plan_type=ee_plan_type).count()
		if CHK>1:
			print ''
		else:
			db =  models.plan_details()
		db.plan_type = ee_plan_type
		db.save()

		return HttpResponse(json.dumps('done'))
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


def delete_entry(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		deletex 	= json.loads(request.GET['delete_id'])
		delete_id 	= models.notification_details.objects.filter(id=deletex).delete()
		
		return HttpResponse(json.dumps('done'))
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')