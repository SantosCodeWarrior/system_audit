from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import WHITE


def inventory_system(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		try:
			max_user_id       = models.equipment_details.objects.filter(active=None).order_by('-label').first()
			max_user_name     = max_user_id.label
			arr 			  = text_num_split(max_user_name)
			flag_number 	  = int(arr[1])+1
			check_length = len(str(flag_number))
			if check_length<=2:
				join_label    = 'BWTW'+'0'+str(flag_number)
			else:
				join_label    = 'BWTW'+str(flag_number)
		except:
			max_user_name = ''
			join_label    = ''

		inventory_details = models.equipment_details.objects.all().order_by('label')		
		inventory_detailx = []

		cursor 			  = connection.cursor()
		Qry				  = "SELECT count(id),category FROM system_audit.audit_equipment_details Group BY category"
		cursor.execute(Qry)
		items_details 	  = cursor.fetchall()
		item_array 		  = []
		for x in items_details:
			item_array.append({
				'total_item' : x[0],
				'item_name'  : x[1]			
				})

		user_details = models.User_entry.objects.all().order_by('user_name')


		d_today = datetime.now().date()+ timedelta(days = 365)
		for v in inventory_details:
			get_loc = models.User_entry.objects.filter(user_name=v.label).first()

			if v.status=='Good':
				color = 'green'
			elif v.status=='Average':
				color = 'orange'
			elif v.status=='Poor':
				color = 'red'
			elif v.status=='Repaired':
				color = 'green'
			elif v.status=='New':
				color = '#1FC71F'
			elif v.status=='Damage':
				color = 'red'
			else:
				color = 'transparent'

			if v.active!=None:
				active_flag = v.active
			else:
				active_flag = 0

			#print '------',v.purchase_date,'----',d_today
			try:
				if v.purchase_date>=d_today:
					expire = 'Expired'
				else:
					expire = ''
			except:
				try:
					if v.repaired_date>=d_today:
						expire = 'Expired'
					else:
						expire = ''
				except:
					pass

			try:
				locations  = get_loc.location
			except:
				locations  = ''

			

			inventory_detailx.append({
				'id'			   : v.id,
				'label' 	       : v.label,
				'brand' 	       : v.brand,
				'status' 	       : v.status,
				'category' 	       : v.category,
				'stock' 	       : v.stock,
				'repaired_date'    : (v.repaired_date),
				'issue_date'       : (v.issue_date),
				'purchase_date'    : (v.purchase_date),
				'remarks'		   : v.remarks,
				'party_name'       : v.party_name,
				'party_contact'    : v.party_contact,
				'mechanic_name'    : v.mechanic_name,
				'mechanic_contact' : v.mechanic_contact,
				'color'			   : color,
				'active'		   : active_flag,
				'expire'		   : expire,
				'locations'		   : locations,
				})

			
		context={
			'max_user_name'     : join_label,
			'inventory_details'	: inventory_detailx,
			'item_array' 		: item_array,
			'user_details'		: user_details
		}	
		return render_to_response("audit_display/inventory_system.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')



def submit_inventory_entry(request):
	category_details			 = request.GET['category_details']
	status_details			  	 = request.GET['status_details']
	brand_details			  	 = request.GET['brand_details']
	max_label			  	 	 = request.GET['max_label']
	stock_details			  	 = request.GET['stock_details']
	purchase_date_details		 = request.GET['purchase_date_details']
	issue_date_details			 = request.GET['issue_date_details']
	repaired_date_details		 = request.GET['repaired_date_details']
	party_name			  	     = request.GET['party_name']
	party_contact			  	 = request.GET['party_contact']
	mechanic_name			  	 = request.GET['mechanic_name']
	mechanic_contact			 = request.GET['mechanic_contact']
	problem_details			  	 = request.GET['problem_details']
	remarks_details			  	 = request.GET['remarks_details']
	active_details 				 = request.GET['active_details']

	max_user_name     = max_label
	arr 			  = text_num_split(max_user_name)
	flag_number 	  = int(arr[1])
	check_length 	  = len(str(flag_number))
	#print '-------',max_label
	if check_length<=2:
		#join_label        = 'BWTW'+'/'+str(category_details)+'/0'+str(flag_number)
		join_label        = 'BWTW'+'/0'+str(flag_number)
	else:
		#join_label        = 'BWTW'+'/'+str(category_details)+'/'+str(flag_number)
		join_label        = 'BWTW'+'/'+str(flag_number)

	check_db = models.equipment_details.objects.filter(label=join_label).count()
	if check_db>0:		
		save_db = models.equipment_details.objects.filter(label=join_label).first()
		msg = 'yes'
	else:	
		save_db = models.equipment_details()
		msg = 'no'
		
	try:
		select_label = category_details.replace(' ','_')
	except:
		select_label = category_details	

	#print '-----',active_details

	save_db.category    		= category_details
	save_db.status		 		= status_details
	save_db.brand    			= brand_details
	save_db.label    			= max_label#join_label
	if active_details!="1":
		save_db.stock    		= stock_details
	else:
		save_db.stock    		= 0
	save_db.active 				= active_details
	
	if purchase_date_details!='':
		save_db.purchase_date   = purchase_date_details
	else:
		save_db.purchase_date   = None

	if issue_date_details!='':
		save_db.issue_date    	= issue_date_details
	else:
		save_db.issue_date    	= None

	if repaired_date_details!='':
		save_db.repaired_date   = repaired_date_details
	else:
		save_db.repaired_date   = None

	save_db.party_name    		= party_name
	save_db.party_contact    	= party_contact
	save_db.mechanic_name    	= mechanic_name
	save_db.mechanic_contact    = mechanic_contact
	save_db.issue_any    		= problem_details
	save_db.remarks    			= remarks_details
	save_db.save()
	return HttpResponse(json.dumps(msg))

def text_num_split(item):
	for index, letter in enumerate(item, 0):
		if letter.isdigit():
			return [item[:index],item[index:]]	

def update_inventory_details(request):
	active 				         = json.loads(request.GET['active'])
	category_details 	         = json.loads(request.GET['category_details'])
	status_details 		         = json.loads(request.GET['status_details'])
	brand_details 		         = json.loads(request.GET['brand_details'])
	stock 				         = json.loads(request.GET['stock'])
	purchase_date 		         = json.loads(request.GET['purchase_date'])
	issue_date 			         = json.loads(request.GET['issue_date'])
	repaired_date 		         = json.loads(request.GET['repaired_date'])
	party_name 			         = json.loads(request.GET['party_name'])
	party_contact 		         = json.loads(request.GET['party_contact'])
	mechanic_name 		         = json.loads(request.GET['mechanic_name'])
	mechanic_contact 	         = json.loads(request.GET['mechanic_contact'])
	problem 			         = json.loads(request.GET['problem'])
	remarks 			         = json.loads(request.GET['remarks'])
	selectID 			         = json.loads(request.GET['selectID'])
	update_db 			         = models.equipment_details.objects.filter(id=selectID).first()	
	#print '---------',purchase_date
	if purchase_date!='':
		update_db.purchase_date	 = purchase_date
	else:
		update_db.purchase_date	 = None

	if issue_date!='':
		update_db.issue_date	 = issue_date
	else:
		update_db.issue_date	 = None

	if repaired_date!='':
		update_db.repaired_date	 = repaired_date
	else:
		update_db.repaired_date	 = None
	
	update_db.active			 = active
	update_db.category	 		 = category_details
	update_db.status	 		 = status_details
	update_db.brand		 		 = brand_details
	try:
		update_db.stock			 = int(stock)
	except:
		update_db.stock			 = 0

	update_db.party_name		 = party_name
	update_db.party_contact		 = party_contact
	update_db.mechanic_name		 = mechanic_name
	update_db.mechanic_contact	 = mechanic_contact
	update_db.issue_any			 = problem
	update_db.remarks			 = remarks
	update_db.save()
	return HttpResponse(json.dumps('done'))


def system_wfh(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		user_details  = models.User_entry.objects.all()	
		user_array = []
		for i in user_details:
			if i.other_items!=None:
				other_items = i.other_items
			else:
				other_items = ''
			user_array.append({
				'id'  			: i.id,			
				'user_name'  	: i.user_name,
				'processor'  	: i.processor,
				'ram'  			: i.ram,
				'hdd'  			: i.hdd,
				'keyboard'      : i.keyboard,
				'mouse'			: i.mouse,
				'smps'			: i.smps,
				'monitor'       : i.monitor,
				'operating_sys' : i.operating_sys,
				'remark'  		: i.remark,
				'items'			: i.items,
				'other_items'	: other_items			
				})

		user_list = models.User_entry.objects.all().order_by('user_name')
		user_details = []
		hp_array     = []
		wfh_array 	 = []
		for x in user_list:
			user_details.append({
				'user_id' : x.id,
				'user_name' : x.user_name
				})

		
		view_wfh_details = models.system_wfh.objects.all()
		for w in view_wfh_details:		
		#	i 			= models.User_entry.objects.filter(id=w.user_id).first()	
			# processor 	= i.processor
			# ram 		= i.ram
			# hdd 		= i.hdd

			wfh_array.append({
				'userID'             : w.id,
				# 'user_name'          : i.user_name,		
				# 'info_hdd'		     : i.hdd,
				# 'info_ram'		     : i.ram,
				# 'info_processor'     : i.processor,
				# 'info_operating_sys' : i.operating_sys,
				# 'info_remarks'  	 : i.remark,
				# 'info_keyboard' 	 : i.keyboard,
				# 'info_mouse'         : i.mouse,
				# 'info_smps' 		 : i.smps,
				# 'info_monitor'       : i.monitor,
				'use_name'			 : w.user_name,
				'keyboard' 			 : w.keyboard,
				'cpu' 				 : w.category,
				'mouse' 			 : w.mouse,
				'ups'				 : w.ups,
				'remarks'			 : w.remarks,
				'issue_date'		 : str(w.issue_date),
				'receive_date' 		 : str(w.receive_date),	

			})
			
		context={
			'user_array'    : user_array,
			'user_details'  : user_details,
			'wfh_array' 	: wfh_array	
		}
		return render_to_response("audit_display/system_wfh.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')

def system_wfh_details(request):
	wfh_array = []
	view_wfh_details = models.system_wfh.objects.all()
	total_wfh 		 = models.system_wfh.objects.filter(receive_date=None).count()

	for w in view_wfh_details:
		wfh_array.append({
			'userID'             : w.id,			
			'use_name'			 : w.user_name,
			'keyboard' 			 : w.keyboard,
			'cpu' 				 : w.category,
			'mouse' 			 : w.mouse,
			'ups'				 : w.ups,
			'remarks'			 : w.remarks,
			'issue_date'		 : str(w.issue_date),
			'receive_date' 		 : str(w.receive_date),	
			'monitor' 			 : w.monitor,
			'headphone'			 : w.headphone,
			'department'		 : w.department,
			'hdmi_cable'		 : w.hdmi_cable,		
		})
		
	context={		
		'wfh_array' : wfh_array,
		'total_wfh' : total_wfh
	}
	return HttpResponse(json.dumps(context))



def submit_wfh_entry(request):
	tabs_name   = json.loads(request.GET['tabs_name'])
	for t in tabs_name:
		if t[1]!=None or t[2]!=None or t[3]!=None or t[4]!=None or t[5]!=None or t[6]!=None  or t[7]!=None or t[8]!=None or t[9]!=None or t[10]!=None or t[11]!=None:
			
			check_DB = models.system_wfh.objects.filter(id=t[0]).count()
			if check_DB>1:
			 	db =  models.system_wfh.objects.filter(id=t[0]).first()
			else:
				db = models.system_wfh()			
			
			ids 		 = t[0]
			user_name    = t[1]
			department   = t[2]
			cpu 	     = t[3]
			keyboard     = t[4]
			mouse 	     = t[5]
			ups 	     = t[6]
			headphone    = t[7]
			monitor      = t[8]
			issue_date   = t[9]
			receive_date = t[10]
			remarks 	 = t[11]		
			
			if receive_date=='None':
				db.receive_date = None
			else:
				db.receive_date = receive_date
			
			if issue_date=='None':
				db.issue_date = None
			else:
				db.issue_date = issue_date

			db.category     	= cpu
			db.keyboard 		= keyboard
			db.mouse 			= mouse				
			db.headphone 		= headphone
			db.remarks 			= remarks
			db.monitor			= monitor
			db.department		= department
			db.user_name 		= user_name
			db.ups 				= ups
			db.id 				= ids
			db.save()

	
	return HttpResponse(json.dumps('done'))


def export_inventory_list(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/inventory_details.xlsx')
	filename 	   					= my_dir+'/inventory_details.xlsx'	
	download_name  					= "Inventory Details.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response





import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

@csrf_exempt
def export_inventory_details(request):
	from shutil import copyfile
	import shutil

	fleet_details = json.loads(request.POST['user_details'])
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/inventory_details.xlsx'
	dstroot       = select_dir+'/inventory_details.xlsx'
	copyfile(srcfile, dstroot)
	my_dir 	      = os.path.dirname(__file__)
	location      = (my_dir+'/inventory_details.xlsx')
	wb 	   	      = load_workbook(my_dir+'/inventory_details.xlsx')
	ws     	      = wb.get_sheet_by_name("sys")
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Arial', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='right'))
	align_style   = Style(font=Font(name='Arial', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='left'))	 	
	file_name     = 'inventory_details.xlsx'

	

	
	i = 3
	j = 1
	for ft in fleet_details:		
		active_name 		= ft['Active']	
		label 	  		    = ft['Label']			
	  	category 	  		= ft['Category']		
	 	status  			= ft['Status']	
	 	brand 		  	  	= ft['Brand']
	 	purchase_date   	= ft["Purchase's Date"]	 	
	 	issue_date 		    = ft["Issue's Date"]
	 	repaired_date 		= ft["Repaired's Date"]
	 	party_name 			= ft["Party's Name"]
	 	party_contact 		= ft["Party's Contact"]
	 	problem 			= ft['Problem']
	 	remarks 		    = ft['Remarks']
	 	location 			= ft['Location']
	 	
	 	
	 	ws.cell('A'+str(i)).value = j	 		
	 	ws.cell('B'+str(i)).value = label
	 	ws.cell('C'+str(i)).value = location	
		ws.cell('D'+str(i)).value = category		
		ws.cell('E'+str(i)).value = status
		ws.cell('F'+str(i)).value = brand
		ws.cell('G'+str(i)).value = purchase_date
		ws.cell('H'+str(i)).value = issue_date
		ws.cell('I'+str(i)).value = repaired_date
		ws.cell('J'+str(i)).value = party_name
		ws.cell('K'+str(i)).value = party_contact
		ws.cell('L'+str(i)).value = problem
		ws.cell('M'+str(i)).value = remarks

		ws.cell('A'+str(i)).style = align_style
		ws.cell('B'+str(i)).style = align_style
		ws.cell('C'+str(i)).style = align_style
		ws.cell('D'+str(i)).style = align_style
		ws.cell('E'+str(i)).style = align_style
		ws.cell('F'+str(i)).style = align_style
		ws.cell('G'+str(i)).style = align_style	
		ws.cell('H'+str(i)).style = align_style	
		ws.cell('I'+str(i)).style = align_style		
		ws.cell('J'+str(i)).style = align_style
		ws.cell('K'+str(i)).style = align_style		
		ws.cell('L'+str(i)).style = align_style
		ws.cell('M'+str(i)).style = align_style
		
		i+=1
		j+=1
	j+=1

	wb.save(my_dir+'/'+file_name)
	return HttpResponse(json.dumps('done'))

