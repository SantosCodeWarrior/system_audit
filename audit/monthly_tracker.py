from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import WHITE


def tracker(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		track_array = []
		_tracker 	= models.monthly_tracker.objects.all()
		for c in _tracker:
			try:
				bill_date = datetime.strptime(str(c.billing_date),"%Y-%m-%d").strftime('%d-%b-%Y')
			except:
				bill_date = ''			

			track_array.append({
				'id' 				: c.id,
				'billing_date' 		: bill_date,
				'service_provider' 	: c.service_provider,
				'staff_member' 		: c.staff_member,
				'bill_no' 			: c.bill_no,
				'nature_expense' 	: c.nature_expense,
				'amount_usd' 		: c.amount_usd,
				'amount_sgd' 		: c.amount_sgd,
				'amount_euro' 		: c.amount_euro,
				'amount_inr' 		: c.amount_inr,
				'total' 			: c.total,
				'remarks' 			: c.remarks,				
			})

		context={
			'track_array' : track_array
		}

		return render_to_response("audit_display/monthly_tracker.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


def submit_tracker_entry(request):
	tabs_name   = json.loads(request.GET['tabs_name'])
	

	for t in tabs_name:
		if t[1]!=None or t[2]!=None or t[3]!=None or t[4]!=None or t[5]!=None or t[6]!=None  or t[7]!=None or t[8]!=None or t[9]!=None or t[10]!=None or t[11]!=None:
			if t[0]!='':
				check_DB = models.monthly_tracker.objects.filter(id=t[0]).count()
				if check_DB>1:
				 	db =  models.monthly_tracker.objects.filter(id=t[0]).first()
				else:
					db = models.monthly_tracker()
			else:
				db = models.monthly_tracker()
			
	 		ids 		 	= t[0]
	 		billing_date 	= t[1]
	 		provider 	 	= t[2]
	 		staff_member 	= t[3]
	 		bill_no	     	= t[4]	
	 		nature_expense 	= t[5]
	 		amount_usd      = t[6]
	 		amount_sgd    	= t[7]
	 		amount_euro     = t[8]
	 		amount_inr 	  	= t[9]
	 		taxable_amt 	= t[10]
	 		total 			= t[11]	 		
	 		remarks 	 	= t[12]
	 		entry_date 		= datetime.now().date()
	 		#print '--------',total

	 		
	 		db.billing_date 	= billing_date			
			db.service_provider	= provider
			db.staff_member 	= staff_member
			db.bill_no 		 	= bill_no
			db.nature_expense 	= nature_expense
			
			
			if amount_usd!='':
				db.amount_usd 	= amount_usd
			else:
				db.amount_usd   = 0
			
			if amount_sgd!='':
				db.amount_sgd 	= amount_sgd
			else:
				db.amount_sgd   = 0
			
			if amount_euro!='':
				db.amount_euro 	= amount_euro
			else:
				db.amount_euro  = 0
			
			if amount_inr!='':
				db.amount_inr 	= amount_inr
			else:
				db.amount_inr   = 0
			
			db.total 		 	= total
			
			try:
				db.remarks 		= remarks
			except:
				db.remarks 		= ''

			try:
				db.taxable      = taxable_amt
			except:
				db.taxable      = 0.0

			db.entry_date  		= entry_date
			db.id 				= ids
	 	 	db.save()

	
	return HttpResponse(json.dumps('done'))


def system_tracker_details(request):
	from django.db.models import Max
	tracker_array 		= []
	view_wfh_details 	= models.monthly_tracker.objects.all()
	last_id 			= models.monthly_tracker.objects.all().aggregate(Max('id'))
	max_id 				= last_id['id__max']
	
	for w in view_wfh_details:
		tracker_array.append({
			'userID'            : w.id,			
			'billing_date'		: str(w.billing_date),
			'service_provider' 	: w.service_provider,
			'staff_member' 		: w.staff_member,
			'bill_no' 			: w.bill_no,
			'nature_expense'	: w.nature_expense,
			'amount_usd'		: w.amount_usd,
			'amount_sgd'		: w.amount_sgd,
			'amount_euro' 		: w.amount_euro,
			'amount_inr' 		: w.amount_inr,
			'total'			 	: w.total,
			'taxable'			: w.taxable,
			'remarks'		 	: w.remarks,
			
		})

		
	context={		
		'tracker_array' : tracker_array,
		'max_id'       	: max_id,		
	}
	
	return HttpResponse(json.dumps(context))