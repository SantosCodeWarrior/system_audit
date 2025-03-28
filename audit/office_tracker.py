from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import WHITE


def office_tracker(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		user_details = models.weekly_entry.objects.filter(entry_date='2024-01-25').order_by('user')
		office365    = []		
		
		for c in user_details:
			total_office3 = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='User3-office-365@bwesglobal.com').count()
			total_office5 = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='User5-office-365@bwesglobal.com').count()
			total_office  = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='bossops@bwesglobal.com').count()
			total_office2 = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='User2.office-365@bwesglobal.com').count()
			total_office7 = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='User7-office-365@bwesglobal.com').count()
			total_office8 = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='User8-office-365@bwesglobal.com').count()
			total_office4 = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='User4-office-365@bwesglobal.com').count()
			total_office6 = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='user6.office@bwesglobal.com').count()
			#total_office1 = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='user.office-365@bwesglobal.com').count()	
			total_account = models.weekly_entry.objects.filter(entry_date=c.entry_date,office365='Accounts').count()		

			office365.append({
				'id' 		 : c.id,
				'offices'    : c.office365,
				'user_name'  : c.user.user_name,
				'department' : c.user.location,
				'assign_to'  : c.user.assigned_to,
				})

		
		context={
			'office365' 	: office365,
			'total_office3' : total_office3,
			'total_office5' : total_office5,
			'total_office'  : total_office ,
			'total_office2' : total_office2,
			'total_office7' : total_office7,
			'total_office8' : total_office8,
			'total_office4' : total_office4,
			'total_office6' : total_office6,
			#'total_office1' : total_office1,
			'total_account' : total_account,

			'subtr_office3' : 10-int(total_office3),
			'subtr_office5' : 10-int(total_office5),
			'subtr_office'  : 10-int(total_office),
			'subtr_office2' : 10-int(total_office2),
			'subtr_office7' : 10-int(total_office7),
			'subtr_office8' : 10-int(total_office8),
			'subtr_office4' : 10-int(total_office4),
			'subtr_office6' : 10-int(total_office6),
			#'subtr_office1' : 10-int(total_office1),
			'subtr_account' : 10-int(total_account),
		}
			
		return render_to_response("audit_display/office_tracker.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')