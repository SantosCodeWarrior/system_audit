from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import WHITE


def system_details(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		user_list = models.User_entry.objects.all().order_by('user_name').exclude(status=1)
		user_details = []
		hp_array     = []
		singel_arr   = []
		today 		 = datetime.now().date()
		for x in user_list:
			get_ids = models.weekly_entry.objects.filter(user_id=x.id,entry_date=today).first()	
			try:
				if get_ids.user.id==x.id:
					display = 'none'
				else:
					display = ''
			except:
				display = ''


			user_details.append({
				'user_id'   : x.id,
				'user_name' : x.user_name,
				'display'   : display
				})

		other_details = models.equipment_details.objects.filter(category='Headphone').order_by('label')	
		for c in other_details:		
			hp_array.append({
				'id'    : c.id,
				'num'   : c.label,			
			})
		context={
		'user_details' : user_details,	
		'hp_array'	   : hp_array,
		}

		print '------',user_details	

		return render_to_response("audit_display/system_details.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')