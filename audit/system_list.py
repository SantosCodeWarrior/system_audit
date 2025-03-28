from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import WHITE


def system_list(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		return render_to_response("audit_display/system_list.html")
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')

@csrf_exempt
def system_details_list(request):
	net_details = models.inventory_list.objects.all()
	net_array 	= []
	#s_no 		= 1
	for x in net_details:
		try:
			reparied_date = (x.repaired_date).strftime('%d-%b-%Y')
		except:
			reparied_date = ''

		try:
			purchase_date = (x.purchase_date).strftime('%d-%b-%Y')
		except:
			purchase_date = ''

		try:
			assigned_date = (x.assigned_date).strftime('%d-%b-%Y')	
		except:
			assigned_date = ''

		if x.repaired_date!=None:
			curr_date 	= datetime.now().date()
			total_day 	= (curr_date-x.repaired_date).days
			#_day 		= str(total_day).split(',')
			get_day 	= str(total_day)+'/365 days'

		else:
			get_day = ''

		net_array.append({
			's_no' 				: x.id,
			'particulars'   	: x.particulars,
			'department'        : x.department,
			'repaired_date'     : reparied_date,
			'purchase_date'     : purchase_date,
			'assigned_to'      	: x.assigned_to,
			'assigned_date'     : assigned_date,
			'status'      		: x.status,
			'warranty'          : x.warranty, #str(get_day),#
			'brand_name'        : x.brand_name,
			'location' 			: x.location,
			'remarks'        	: x.remarks,
			
			})
		#s_no+=1

	context={
	'net_array'  : net_array,
	}	
	
	return HttpResponse(json.dumps(context))


@csrf_exempt
def system_details_entry(request):	

	tabs_name=json.loads(request.POST['tabs_name'])	
	for t in tabs_name:		
		check = models.inventory_list.objects.filter(id=t[0]).count()
		if check==1:
			net = models.inventory_list.objects.filter(id=t[0]).first()
		else:		
			net = models.inventory_list()

		if t[1]!=None or t[2]!=None or t[3]!=None or t[4]!=None or t[5]!=None or t[6]!=None  or t[7]!=None or t[8]!=None or t[9]!=None or  t[10]!=None or  t[11]!=None:			
			try:
				reparied_date = datetime.strptime(t[3], "%d-%b-%Y").strftime('%Y-%m-%d')
			except:
				reparied_date = None
			
			try:
				purchase_date = datetime.strptime(t[4], "%d-%b-%Y").strftime('%Y-%m-%d')
			except:
				purchase_date = None

			try:
				assigned_date = datetime.strptime(t[6], "%d-%b-%Y").strftime('%Y-%m-%d')
			except:
				assigned_date = None

			net.particulars 			= t[1]
			net.department 				= t[2]

			try:
				net.repaired_date 		= reparied_date
			except:
				net.repaired_date  		= None			
			
			try:
				net.purchase_date 		= purchase_date	
			except:
				net.purchase_date 		= None			
			
			try:
				net.assigned_date 		= assigned_date	
			except:
				net.assigned_date 		= None

			net.assigned_to 			= t[5]	
			net.status  				= t[7]
			net.warranty 				= t[8]
			net.brand_name 				= t[9]
			net.location 				= t[10]
			net.remarks 				= t[11]				
		 	net.save()
	
	return HttpResponse(json.dumps("done"))










