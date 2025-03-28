from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum

def view_audit_summary(request):
	recorded_date   = '2021-07-03'
	total_pc   		= models.User_entry.objects.filter(items='Desktop').count()
	total_dp   		= models.User_entry.objects.filter(items='DP Machine').count()	
	tot_opr    		= models.weekly_entry.objects.filter(use_system='Operator',entry_date=recorded_date).count()
	tot_master 		= models.weekly_entry.objects.filter(use_system='Master',entry_date=recorded_date).count()
	tot_chm    		= models.weekly_entry.objects.filter(use_system='CHM',entry_date=recorded_date,user__items='Desktop').count()
	tot_it     		= models.weekly_entry.objects.filter(use_system='IT',entry_date=recorded_date,user__items='Desktop').count()
	tot_recp   		= models.weekly_entry.objects.filter(entry_date=recorded_date,use_system='Reception').count()
	damage_details	= models.equipment_details.objects.all()
	i=0
	for c in damage_details:
		if c.status=='Poor':			
			print '------',c.label
			i+=1

	context={
		'total_pc'   : total_pc,
		'total_dp'   : total_dp,
		'tot_opr'    : tot_opr,
		'tot_master' : tot_master,
		'tot_chm'    : tot_chm,
		'tot_it'     : tot_it,
		'tot_recp'   : tot_recp,
		'tot_dage'   : 'tot_dage',
		'tot_poor'   : 'tot_poor'
	}

	return render_to_response("audit_display/view_audit_summary.html",context)