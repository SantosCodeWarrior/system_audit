from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range

try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt

from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import WHITE


def view_bills(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':

		return render_to_response("audit_display/view_bills.html")
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


def get_bills_details(request):
	bills 		 = request.GET['bills']
	bills_list 	 = models.equipment_details.objects.filter(id=bills).first()
	get_location = models.User_entry.objects.filter(user_name=bills_list.label).first()
	if bills_list.active==1:
		statuss  = 'Active'
	else:
		statuss  = 'Not working'
	
	context={
		'system_id'  	: bills_list.label,
		'category'   	: bills_list.category,
		'brand' 	 	: bills_list.brand,
		'party_name' 	: bills_list.party_name,
		'status'	 	: statuss,
		'purchase_date' : str(bills_list.purchase_date),
		'repaired_date' : str(bills_list.repaired_date),
		'location'		: get_location.location,
		'get_id' 		: get_location.id,
	}

	return HttpResponse(json.dumps(context))

def submit_bills_details(request):
	e_gst			= json.loads(request.GET['e_gst'])
	sys_id			= json.loads(request.GET['sys_id'])
	e_sgst			= json.loads(request.GET['e_sgst'])
	e_total			= json.loads(request.GET['e_total'])
	e_g_total		= json.loads(request.GET['e_g_total'])
	e_bill_date		= json.loads(request.GET['e_bill_date'])
	e_label_name	= json.loads(request.GET['e_label_name'])
	e_warranty_time	= json.loads(request.GET['e_warranty_time'])
	



	return HttpResponse(json.dumps('done'))