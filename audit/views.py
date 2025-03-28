from django.shortcuts import render
from django.shortcuts import render,render_to_response
from audit import models
from django.http import HttpResponse
import json
from django.db.models import Count
from django.db import connection
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import calendar

import smtplib
import pprint
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives
from email.MIMEImage import MIMEImage

import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from django.db.models import Sum


def dashboard(request):
	print ''
	return render_to_response("base.html",{})


def welcome(request):	
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		day_of_week = 5
		max_weeks   = 43022
		models.weekly_name_details.objects.all().delete()
		for d in get_all_dates_of_day_of_week_in_year (day_of_week, 2019, 1, 1, max_weeks):
			cal 		=  "%s, %d/%d/%d" % (calendar.day_name[d.weekday()], d.year, d.month, d.day)
			split_cal 	= cal.split(',')
			date_filter = (split_cal[1])
			if d.month<10:
				fin_month = '0'+str(d.month)
			else:
				fin_month = d.month

			if d.day<10:
				fin_day  = '0'+str(d.day)
			else:
				fin_day = d.day

			current_date = datetime.now().date()
			chk_month    = current_date.strftime("%m")
			chk_day		 = current_date.strftime("%d")
			
			if str(chk_month)==str(fin_month):
				check =  models.weekly_name_details.objects.filter(week_name=fin_month,date=d.day).count()
				if check>1:
					wkly = models.weekly_name_details.objects.filter(week_name=fin_month,date=d.day).first()
				else:
					wkly  = models.weekly_name_details()
				wkly.week_name = fin_month
				wkly.date      = d.day
				wkly.save()

				cnd = models.weekly_name_details.objects.filter(week_name=fin_month,date=d.day).first()
				if str(chk_month)==str(cnd.week_name):
					if chk_day==cnd.date:
						next_audit = 'Audit : Sat, '+str(d.day)
					else:
						next_audit = ''



		# bcc_email 	   	= 'santosh@bwesglobal.com'
		# select_dir     	= os.path.dirname(__file__)
		# curr_date 	   	= datetime.now().date()
		# remind 			= models.remainder_details.objects.filter(office_period_date=curr_date).first()
		# if remind!=None:
		# 	html_content   = 'Good day Sir,<br><br><br>Greetings from Team Enet,<br><br>Please Recharge for your Internet Package due on <b>'+str(curr_date)+ '</b><br><br>Thanks & Best Regards<br><b style="color:grey;font-size:12px">SYSTEM AUDIT</b>'	
		# 	text_content   = ''		
		# 	mailto 		   = 'santosh@bwesglobal.com'
		# 	cc_email 	   = 'santosh@bwesglobal.com'
		# 	mailsubject    = 'Recharge for your Internet Package (E-Net)'	
		# 	msg  		   = EmailMultiAlternatives(mailsubject,text_content,'santosh@bwesglobal.com',[mailto],bcc=[bcc_email],cc=[cc_email])		
		# 	msg.attach_alternative(html_content, "text/html")
			#msg.send()

		check_list = models.weekly_entry.objects.filter(entry_date=datetime.now().date()).count()
		#print '-----',check_list
		if check_list>1:
			show = ''
		else:
			show = ''

		final_arr 	= []
		single_arr 	= []
		status 		= ''
		# curr_user = models.weekly_entry.objects.all()		
		curr_user = models.User_entry.objects.all().order_by('user_name')
		i=0
		for v in curr_user:	
		 	try:
			 	check_name = models.weekly_entry.objects.filter(entry_date=datetime.now().date(),user__user_name=v.user_name).first()
			 	if check_name.user.user_name==v.user_name:
			 		icon   = 'fa fa-check'
			 		color  = 'green'
			 		status = 'none'
			 	else:
			 		icon   = 'fa fa-times'
			 		color  = 'red'
			 		status = ''
				#print '---------',status,'----',v.user_name
			except:
			 	icon   = 'fa fa-times'
			 	color  = 'red'
			 	status = ''

			try:
				loca = v.location
			except:
				loca = ''

			if v.location=="Store Room" or v.location=="WFH":
				color_Code = "red"
			else:
				color_Code = "green"

			print '----------',v.location
			

		 	final_arr.append({
		 		'user_name' : v.user_name,	
		 		'icon'      : icon,	
		 		'color' 	: color, 
		 		'status'    : status,  
		 		's_no'		: (i+1),
		 		'sid'  		: v.id,
		 		'location'  : loca,
		 		'colorCode' : color_Code
		 		
		 		})
		 	i+=1
				
		
			
		context={
			'next_audit' : next_audit,		
			'show'       : show,
			'final_arr'  : final_arr,
		}	
		return render_to_response("audit_display/welcome.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')

def Remove(duplicate): 
	final_list = [] 
	for num in duplicate: 
		if num not in final_list: 
			final_list.append(num) 
	return final_list 

def get_all_dates_of_day_of_week_in_year(day_of_week, start_year, start_month,start_day, max_weeks=None):
    if day_of_week < 0 or day_of_week > 6:
        raise ValueError('day_of_week should be in [0, 6]')
    date_iter = date(start_year, start_month, start_day)    
    date_iter += timedelta(days=(day_of_week - date_iter.weekday() + 7) % 7) 
    week = 1
    while date_iter.year == start_year:
        yield date_iter
        date_iter += timedelta(days=7)
        if max_weeks is not None:
            week += 1
            if week > max_weeks:
                break

def user_entry(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		#if 'voy_id' in request.GET:
		user_details  = models.User_entry.objects.all().exclude(status=1).order_by('user_name')
		other_details = models.equipment_details.objects.filter(active=None).order_by('label')
		
		hp_array  	  = []
		for c in other_details:		
			hp_array.append({
				'id'    : c.id,
				'num'   : c.label,
				#'stock' : stock
			})

		user_array = []
		for i in user_details:
			if i.other_items!=None:
				other_items = i.other_items
			else:
				other_items = ''

			if i.location!=None:
				loc = i.location
			else:
				loc = ''

			if i.assigned_to!=None:
				assigned_to = i.assigned_to
			else:
				assigned_to = ''

			#print '--------',assigned_to




			# get_location  = models.weekly_entry.objects.filter(user_id=i.id).first()
			
			# try:			
			# 	if get_location.depart_name!=None:
			# 		location_name = get_location.depart_name

			# 	else:
			# 		location_name = ""
			# except:
			# 	location_name = ''



			

			user_array.append({
				'id'  			: i.id,			
				'user_name'  	: i.user_name,
				'processor'  	: i.processor,
				'ram'  			: i.ram,
				'hdd'  			: i.hdd,
				'keyboard'      : i.keyboard,
				'mouse'			: i.mouse,
				'smps'			: i.smps,
				'monitor'       : i.monitor,
				'operating_sys' : i.operating_sys,
				'remark'  		: i.remark,
				'items'			: i.items,
				'other_items'	: other_items,
				'location'		: loc,	
				'get_assigned'  : assigned_to,		
				})
		context={
			'user_array'    : user_array,
			'hp_array'      : hp_array
		}
		
		return render_to_response("audit_display/user_entry.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')

	

def submit_entry(request):	
	user_name 		= request.GET['user']
	processor   	= request.GET['processor']
	ram 			= request.GET['ram']
	hdd 			= request.GET['hdd']
	ope_sys  		= request.GET['ope_sys']
	remark  		= request.GET['remark']	
	headphone_key   = request.GET['headphone_key']
	other_details	= request.GET['computer_details']

	user = ''
	user_already    = models.User_entry.objects.filter(user_name=user_name).count()	
	if user_already==1:		
		msg = 'already'
	else:
		user 			  = models.User_entry()
		msg = 'done'

	user.user_name 		 = user_name
	user.processor  	  = processor
	user.ram 			  = ram
	user.hdd 			  = hdd
	user.operating_sys    = ope_sys
	user.remark 		  = remark
	user.items 			  = other_details
	user.other_items      = headphone_key	
	user.save()
	# other_entry 		  = models.equipment_details.objects.filter(label=headphone_key).first()
	# try:
	# 	other_entry.active 	  = 1
	# except:
	# 	other_entry.active 	  = 0
	# other_entry.save()
	return HttpResponse(json.dumps(msg))

def entry_form(request):	
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		user_list = models.User_entry.objects.filter(active=None).order_by('user_name')
		user_details = []
		hp_array     = []
		singel_arr   = []
		today 		 = datetime.now().date()
		for x in user_list:
			get_ids = models.weekly_entry.objects.filter(user_id=x.id,entry_date=today).first()	
			try:
				if get_ids.user.id==x.id:
					display = 'none'
				else:
					display = ''
			except:
				display = ''

			try:
				user_name = str(x.user_name)
			except:
				user_name = ''


			user_details.append({
				'user_id'   : x.id,
				'user_name' : user_name,
				'display'   : display
				})

		other_details = models.equipment_details.objects.filter(category='Headphone').order_by('label')	
		for c in other_details:		
			hp_array.append({
				'id'    : c.id,
				'num'   : c.label,			
			})
		context={
		'user_details' : user_details,	
		'hp_array'	   : hp_array,
		}		
		return render_to_response("audit_display/entry.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


def submit_weekly_entry(request):	
	user			 = request.GET['user']	
	boot_time		 = request.GET['boot_time']
	unused_program	 = request.GET['unused_program']
	temp_file		 = request.GET['temp_file']
	latest_antivirus = request.GET['latest_antivirus']
	history_clean	 = request.GET['history_clean']
	server_conn		 = request.GET['server_conn']
	internet_conn	 = request.GET['internet_conn']
	printer_conn	 = request.GET['printer_conn']
	head_phone  	 = request.GET['head_phone']
	issue_any  		 = request.GET['issue_any']
	head_phone_num   = request.GET['head_phone_num']

	depart_name      = request.GET['depart_name']
	use_system       = request.GET['use_system']
	email_configure  = request.GET['email_configure']
	assigned_to 	 = request.GET['assigned_to']
	office365 		 = request.GET['office365']

	check_date 		 = datetime.now().date()
	entry_array 	 = []
	update_date      = models.User_entry.objects.filter(id=user).first()
	check_entry 	 = models.weekly_entry.objects.filter(user_id=user,entry_date=check_date).count()	
	
	if check_entry==1:
		entry = models.weekly_entry.objects.filter(user_id=user,entry_date=check_date).first()		
		msg   = 'repeat'
		return HttpResponse(json.dumps(msg))
	else:
		entry = models.weekly_entry()	
	entry.entry_date 		= datetime.now().date()
	entry.user_id 			= user
	entry.boot_time 		= boot_time 	 
	entry.unused_program	= unused_program	 
	entry.temp_file		   	= temp_file	 
	entry.up_antivirus      = latest_antivirus		 
	entry.history_clean	  	= history_clean
	entry.server_folder	  	= server_conn	 
	entry.internet_conn	  	= internet_conn		
	entry.printer_conn	  	= printer_conn	 
	entry.issue_any  	  	= issue_any
	entry.head_phone 		= head_phone
	entry.depart_name       = depart_name
	entry.use_system      	= use_system
	entry.email_configure   = email_configure
	entry.current_date      = datetime.now().date()
	entry.other_items       = head_phone_num
	entry.assigned_to 		= assigned_to
	entry.office365 		= office365

	entry.save()
	msg = 'done'

	submit_date     		 = models.User_entry.objects.filter(id=user).first()
	submit_date.current_date = datetime.now().date()
	submit_date.other_items  = head_phone_num
	submit_date.save()


	return HttpResponse(json.dumps(msg))

def refresh_weekly_entry(request):
	current_date  = datetime.now().date()
	chk_month     = current_date.strftime("%m")
	chk_day		  = current_date.strftime("%d")

	check_details = models.weekly_name_details.objects.all()
	for e in check_details:	
		if e.date!=chk_day:	
			clean_current_entry = models.User_entry.objects.all()
			for c in clean_current_entry:
				c.current_date = None			
				c.save()

	return HttpResponse(json.dumps('done'))


def remainder(request):
	user_list    = models.User_entry.objects.all().order_by('user_name')
	user_details = []
	for x in user_list:		
		user_details.append({
			'user_id'   : x.id,
			'user_name' : x.user_name
			})

	enail_list    = models.email_details.objects.all().order_by('email_user_name')
	email_details = []
	for xx in enail_list:		
		email_details.append({
			'userID'     : xx.id,
			'email_name' : xx.email_user_name
			})

	key_details = models.remainder_details.objects.all().order_by('id')
	key_array = []
	for k in key_details:
		try:
			pic_name = k.user.user_name
		except:
			pic_name = ''
		key_array.append({
			'id' 			 : k.id,
			'remainder_date' : str(k.remainder_date),
			'user_name'		 : k.user_name,
			'pic_name'       : pic_name,
			'remark' 		 : k.remarks,
			'mail_moved_date': str(k.office_period_date),
			'userID' 		 : k.user_id,
			})

	context={
	'user_details'  : user_details,
	'email_details' : email_details,
	'key_array'     : key_array,
	}

	return render_to_response("audit_display/remainder.html",context)


def view_audit(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		equip_details = models.equipment_details.objects.all().exclude(category='CPU').exclude(category='Desktop').exclude(category='Headphone').exclude(category='UPS').exclude(category='Keyboard').exclude(category='Mouse').exclude(category='Monitor').exclude(category='Mouse Pad').exclude(category='RAM').order_by('category')		
		equip_array   = []
		for c in equip_details:
			if c.status=='Good':
				iconx   = "fas fa-thumbs-up"
				bgcolor = "green"
			elif c.status=='Poor':
				iconx   = 'fas fa-thumbs-down'
				bgcolor = "red"
			elif c.status=='Damaged':
				iconx   = 'fas fa-thumbs-down'
				bgcolor = "red"
			else:
				iconx   = ''
				bgcolor = 'black'

			equip_array.append({
				'items'   : c.category,
				'label'   : c.label,
				'status'  : c.status,
				'iconx'   : iconx,
				'bgcolor' : bgcolor,
				})
		context={
		'equip_array' : equip_array
		}
		return render_to_response("audit_display/view_audit.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


def selected_date_view_audit(request):
	selected_date  			= request.GET['selected_date']
	current_date   			= datetime.now().date()
	chk_year  	   			= current_date.strftime("%Y")
	if selected_date:	
		selected_entry_date = datetime.strptime(str(selected_date),"%d/%b/%Y").strftime('%Y-%m-%d')
	else:
		max_date 			= models.weekly_entry.objects.all().order_by('-entry_date').first()		
		selected_entry_date = str(max_date.entry_date)	
	
	weekly_details 			= models.weekly_entry.objects.filter(entry_date=selected_entry_date).order_by('user__user_name')
	weekly_array   			= []
	for i in weekly_details:
		for_date = (i.entry_date).strftime('%d-%b-%Y')
		if i.boot_time=='':
			bg_color = '#f57a7a'
		else:
			bg_color = ''

		
		weekly_array.append({
			'userID'             : i.id,
			'user_name'          : i.user.user_name,
			'boot_time'          : i.boot_time,
			'entry_date'         : str(for_date),
			'unused_program'     : i.unused_program,
			'temp_file'          : i.temp_file,	
			'up_antivirus'       : i.up_antivirus,
			'history_clean'      : i.history_clean,
			'server_folder'      : i.server_folder,
			'internet_conn'      : i.internet_conn,
			'printer_conn'       : i.printer_conn,
			'issue_any'          : i.issue_any,
			'department' 		 : i.depart_name,
			'use_system'		 : i.use_system,
			'email_configure'	 : i.email_configure,
			'user_id'            : i.user_id,
			'info_hdd'		     : i.user.hdd,
			'info_ram'		     : i.user.ram,
			'info_processor'     : i.user.processor,
			'info_operating_sys' : i.user.operating_sys,
			'info_remarks'  	 : i.user.remark,
			'info_keyboard' 	 : i.user.keyboard,
			'info_mouse'         : i.user.mouse,
			'info_smps' 		 : i.user.smps,
			'info_monitor'       : i.user.monitor,
			'head_phone'		 : i.head_phone,
			'other_items'  		 : i.other_items,
			'bg_color' 			 : bg_color,
			'assigned_to'		 : i.assigned_to,
			#'mail_color'         : mail_color,
		})

	context={
		'weekly_array'  : weekly_array
	}

	return HttpResponse(json.dumps(context))

def submit_remainder_entry(request):	
	user_name   	  	 = request.GET['user_name']
	pic_name   		  	 = request.GET['pic_name']
	e_moved_email_date   = request.GET['e_moved_email_date']
	e_period_date     	 = request.GET['e_period_date']
	e_remark   		  	 = request.GET['e_remark']
	audit_arr 		  	 = models.User_entry.objects.filter(user_name=pic_name).first()	
	check 			  	 = models.remainder_details.objects.filter(user_id=str(audit_arr.id)).count()
	
	if check>1:
		save_db = models.remainder_details.objects.filter(user_id=audit_arr.id).first()
		msg = 'no'		
	else:
		save_db = models.remainder_details()
		msg = 'done'
		
	save_db.user_id  			=  audit_arr.id
	save_db.remainder_date 		= datetime.now().date()
	save_db.office_period_date 	= e_period_date
	save_db.email_move_date   	= e_moved_email_date
	save_db.remarks 		   	= e_remark
	save_db.user_name 			= user_name
	save_db.save()
	#print '==========',msg

	context={
	'msg'  : msg
	}

	return HttpResponse(json.dumps(context))

def assets_register(request):
	user_list    = models.User_entry.objects.all().order_by('user_name')
	user_details = []
	for x in user_list:		
		user_details.append({
			'user_id'   : x.id,
			'user_name' : x.user_name
			})

	try:
		max_key 	= models.register_details.objects.all().order_by('-audit_key').first()
		next_arr 	= max_key.audit_key
		first_array = next_arr.split('/')
		add 		= int(first_array[2])+1
		next_key = 'BW/CPU/00'+str(add)
	except:
		next_key =  'BW/CPU/001'


		
	context={
		'user_details' : user_details,
		'next_key'     : next_key
	}
	return render_to_response("audit_display/assets_register.html",context)


def submit_register_entry(request):
	e_id     		  = json.loads(request.GET['e_id'])
	e_non_it          = json.loads(request.GET['e_non_it'])
	e_fitting         = json.loads(request.GET['e_fitting'])
	e_description     = json.loads(request.GET['e_description'])
	e_location 		  = json.loads(request.GET['e_location'])
	e_user_allotted   = json.loads(request.GET['e_user_allotted'])
	e_security_level  = json.loads(request.GET['e_security_level'])
	e_maintenance 	  = json.loads(request.GET['e_maintenance'])
	e_record_date	  = json.loads(request.GET['e_record_date'])
	e_remarks 		  = json.loads(request.GET['e_remarks'])

	try:
		max_key = models.register_details.objects.all().order_by('-audit_key').first()		
	except:
		pass

	audit_arr = models.User_entry.objects.filter(user_name=e_user_allotted).first()	
	check 	  = models.register_details.objects.filter(user_id=audit_arr.id,user_allotted=e_user_allotted).count()
	if check>1:
		inser_db = models.register_details.objects.filter(user_id=audit_arr.id,user_allotted=e_user_allotted).first()		
	else:
		inser_db = models.register_details()

		inser_db.audit_key      = e_id
		inser_db.it_non 		= e_non_it
		inser_db.fitting 		= e_fitting
		inser_db.description 	= e_description
		inser_db.location  		= str(e_location)
		inser_db.user_allotted 	= e_user_allotted
		inser_db.security_level = e_security_level
		inser_db.maintenance 	= e_maintenance
		inser_db.user_id 		= audit_arr.id
		inser_db.record_date    = datetime.now().date()
		if e_record_date:
			inser_db.register_date 	= e_record_date
		else:
			inser_db.register_date = None
		inser_db.remarks       	= e_remarks
		inser_db.save()
	
	return HttpResponse(json.dumps('done'))

def view_register_entry(request):
	register_details = models.register_details.objects.all()
	reg_array = []	
	for x in register_details:
		reg_array.append({
			'audit_key'		: x.audit_key,
			'it_non'		: x.it_non,
			'fitting'		: x.fitting,
			'description'	: x.description,
			'location'		: x.location,
			'user_allotted'	: x.user_allotted,
			'security_level': x.security_level,
			'maintenance'	: x.maintenance,
			'record_date'	: str(x.record_date),
			'remark'        : x.remarks,
			'register_date' : str(x.register_date),
			'userID'        : x.user.id
			}) 
	context={
		'reg_array'  : reg_array
	}
	return render_to_response("audit_display/view_assets_register.html",context)




def update_user_entry(request):	
	user		= json.loads(request.GET['user'])
	processor	= json.loads(request.GET['processor'])
	ram			= json.loads(request.GET['ram'])
	hdd			= json.loads(request.GET['hdd'])
	ope_sys		= json.loads(request.GET['ope_sys'])
	remark		= json.loads(request.GET['remark'])

	keyboard    = json.loads(request.GET['keyboard'])
	mouse       = json.loads(request.GET['mouse'])
	monitor     = json.loads(request.GET['monitor'])
	smps        = json.loads(request.GET['smps'])	
	items_types = json.loads(request.GET['items_types'])
	headphone_k = json.loads(request.GET['headphone_key'])
	depart_name = json.loads(request.GET['depart_name'])
	assigned_to = json.loads(request.GET['assigned_toz'])

	other_entry = 0
	update_user_entry = models.User_entry.objects.filter(user_name=user).first()
	update_user_entry.user 		 	 = user
	update_user_entry.processor  	 = processor
	update_user_entry.ram 		 	 = ram
	update_user_entry.hdd 		 	 = hdd
	update_user_entry.operating_sys  = ope_sys
	update_user_entry.remark 	     = remark
	update_user_entry.keyboard 	     = keyboard
	update_user_entry.mouse 	     = mouse
	update_user_entry.monitor 	     = monitor
	update_user_entry.smps 	     	 = smps
	update_user_entry.items 		 = items_types
	update_user_entry.location 		 = depart_name
	try:
		update_user_entry.assigned_to  = assigned_to
	except:
		update_user_entry.assigned_to  = None

	try:
		update_user_entry.other_items = headphone_k
	except:
		update_user_entry.other_items = None
	
	update_user_entry.save()

	try:
		other_entry 		= models.equipment_details.objects.filter(label=headphone_k).first()	
		other_entry.active 	= 1
		other_entry.save()
	except:	
		pass

	return HttpResponse(json.dumps('done'))

def view_system_configuration(request):
	userID 		   = json.loads(request.GET['userID'])
	weekly_details = models.User_entry.objects.filter(id=userID)
	weekly_array   = []
	for i in weekly_details:
		if i.other_items!=None and i.other_items!="":
			other_itemxs = i.other_items
		else:
			other_itemxs = ''

		if i.items=='Desktop':
			img   = '/static/img/pc.png'
		elif i.items=='Reception':
			img   = '/static/img/reception.jpg'			
		elif i.items=='Laptop':
			img   = '/static/img/laptop.png'			
		elif i.items=='DP Machine':
			img   = '/static/img/dp.png'
		elif i.items=='Server':
			img   = '/static/img/server.jpg'	
		elif i.items=='Camera':
			img   = '/static/img/camera.png' 
		elif i.items=='CCTV':
			img   = '/static/img/cctv.png' 				
		else:
			img   = '/static/img/blank.png'

		if i.other_items!=None and i.other_items!="#" and i.other_items!="":
			k_img = '/static/img/headphone.jpg'
		else:	
			k_img = '/static/img/blank.png'

		try:
			assigned_to = i.assigned_to
		except:
			assigned_to = ''


		
		weekly_array.append({
			'userID'             : i.id,
			'user_name'          : i.user_name,			
			'user_id'            : i.id,
			'info_hdd'		     : i.hdd,
			'info_ram'		     : i.ram,
			'info_processor'     : i.processor,
			'info_operating_sys' : i.operating_sys,
			'info_remarks'  	 : i.remark,
			'info_keyboard' 	 : i.keyboard,
			'info_mouse'         : i.mouse,
			'info_smps' 		 : i.smps,
			'info_monitor'       : i.monitor,
			'img' 				 : img,
			'k_img' 			 : k_img,
			'serial_no'			 : other_itemxs,
			'assgined_for'		 : assigned_to,
		})

	context={
		'weekly_array'  : weekly_array
	}
	
	
	return HttpResponse(json.dumps(context))

def view_user_configuration(request):
	userID = json.loads(request.GET['userID'])	
	weekly_details = models.User_entry.objects.filter(id=userID)
	weekly_array   = []
	for i in weekly_details:		
		weekly_array.append({
			'userID'             : i.id,
			'user_name'          : i.user_name,		
			'info_hdd'		     : i.hdd,
			'info_ram'		     : i.ram,
			'info_processor'     : i.processor,
			'info_operating_sys' : i.operating_sys,
			'info_remarks'  	 : i.remark,
			'info_keyboard' 	 : i.keyboard,
			'info_mouse'         : i.mouse,
			'info_smps' 		 : i.smps,
			'info_monitor'       : i.monitor,
		})

	context={
		'weekly_array'  : weekly_array
	}
	
	return HttpResponse(json.dumps(context))

def network_speed_details(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		return render_to_response("audit_display/network_record.html")
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


@csrf_exempt
def misc_details(request):
	net_details = models.network_details.objects.all()
	net_array = []
	s_no= 1
	for x in net_details:	
		
		net_array.append({
			's_no' 					: x.id,
			'break_down'            : str(x.break_down),
			'action_date'           : str(x.action_date),
			'name_mc'               : x.name_mc,
			'reported_name'         : x.reported_name,
			'network_operator'      : x.network_operator,
			'problem_reported'      : x.problem_reported,
			'action_taken'          : x.action_taken,
			'time_status'           : str(x.time_status),
			'total_time_break_down' : x.total_time_break_down,
			'internet_speed'        : x.internet_speed,
			'remarks'               : x.remarks,
			})
		s_no+=1
	context={
	'net_array'  : net_array,
	}	
	
	return HttpResponse(json.dumps(context))

@csrf_exempt
def misc_details_entry(request):	
	tabs_name=json.loads(request.POST['tabs_name'])	
	for t in tabs_name:	
		
		check = models.network_details.objects.filter(id=t[0]).count()
		if check==1:
			net = models.network_details.objects.filter(id=t[0]).first()
		else:		
			net = models.network_details()

		if t[1]!=None or t[2]!=None or t[3]!=None or t[4]!=None or t[5]!=None or t[6]!=None  or t[7]!=None or t[8]!=None or t[9]!=None or  t[10]!=None:			
			#print '===========',t[0]
			net.break_down 				= t[1][0:19]
			net.name_mc 				= t[2]
			net.reported_name 			= t[3]
			net.network_operator 		= t[4]
			net.problem_reported 		= t[5]
			net.action_taken 			= t[6]
			net.time_status 			= t[7][0:19]
			net.total_time_break_down 	= t[8]
			net.internet_speed 			= t[9]
			net.remarks 				= t[10]	
			net.action_date 			= datetime.now().date()		
		 	net.save()
	
	return HttpResponse(json.dumps("done"))
	

def system_log(request):
	user_list    = models.User_entry.objects.all().order_by('user_name')
	user_details = []
	for x in user_list:
		user_details.append({
			'user_id'   : x.id,
			'user_name' : x.user_name
			})
	log_detaila = models.system_problem_details.objects.all()
	log_arr =[]
	for k in log_detaila:
		try:
			user_name = k.user.user_name
			userID    = k.user.id
		except:
			user_name = ''
			userID    = ''

		log_arr.append({
			'id' 			: k.id,
			'party_name'    : k.party_name,
			'party_number'  : k.party_contact,
			'mechanic_name' : k.mechanic_name,
			'mechanic_no'   : k.mechanic_number,
			'problem'       : k.problems,
			'user_name'     : user_name,
			'remarks'       : k.remarks,
			'action_date'   : k.called_date.strftime('%d-%b-%Y'),
			'userID'    	: userID,
 			})

	context={
	'user_details' : user_details,
	'log_arr'      : log_arr,	
	} 	
	return render_to_response("audit_display/system_log.html",context)

def submit_log(request):
	e_user_name         = json.loads(request.GET['e_user_name'])
	e_party_name        = json.loads(request.GET['e_party_name'])
	e_party_number      = json.loads(request.GET['e_party_number'])
	e_mechanic_name     = json.loads(request.GET['e_mechanic_name'])
	e_mechanic_number   = json.loads(request.GET['e_mechanic_number'])
	e_problem           = json.loads(request.GET['e_problem'])
	e_remarks           = json.loads(request.GET['e_remarks'])
	
	user_det = models.User_entry.objects.filter(user_name=e_user_name).first()	
	check =  models.system_problem_details.objects.filter(user_id=user_det.id).count()
	# if check==1:
	# 	msg = 'no'
	# 	return HttpResponse(json.dumps(msg))
	# else:
	e_log 	 = models.system_problem_details()
	msg = 'done'
	e_log.user_id 			=  user_det.id
	e_log.party_name 		=  e_party_name
	e_log.party_number 		=  e_party_number
	e_log.mechanic_name 	=  e_mechanic_name
	e_log.mechanic_number 	=  e_mechanic_number
	e_log.problem 			=  e_problem
	e_log.remarks 			=  e_remarks
	e_log.called_date 		= datetime.now().date()
	e_log.save()


	return HttpResponse(json.dumps(msg))

def view_sys_info(request):
	userID = json.loads(request.GET['user'])
	#print '===============USERID',userID
	weekly_details = models.User_entry.objects.filter(id=userID)
	weekly_array   = []
	for i in weekly_details:		
		weekly_array.append({
			'userID'             : i.id,
			'user_name'          : i.user_name,		
			'info_hdd'		     : i.hdd,
			'info_ram'		     : i.ram,
			'info_processor'     : i.processor,
			'info_operating_sys' : i.operating_sys,
			'info_remarks'  	 : i.remark
		})

	context={
	'weekly_array'  : weekly_array
	}
	return HttpResponse(json.dumps(context))

def update_sys_log(request):
	logID 			= json.loads(request.GET['_userID'])
	_party_name     = json.loads(request.GET['_party_name'])
	_party_no     	= json.loads(request.GET['_party_no'])
	_mechanic_name  = json.loads(request.GET['_mechanic_name'])
	_mechanic_no    = json.loads(request.GET['_mechanic_no'])
	_ex_problems    = json.loads(request.GET['_ex_problems'])
	_ex_remarks     = json.loads(request.GET['_ex_remarks'])

	log_update = models.system_problem_details.objects.filter(id=logID).first()
	log_update.party_name 		= _party_name
	log_update.party_contact  	= _party_no
	log_update.mechanic_name  	=  _mechanic_name
	log_update.mechanic_number  = _mechanic_no
	log_update.problems         = _ex_problems
	log_update.remarks 			= _ex_remarks
	log_update.save()
	return HttpResponse(json.dumps('done'))


def update_key_details(request):
	ex_selectID  = json.loads(request.GET['ex_selectID'])
	ex_user_name = json.loads(request.GET['ex_user_name'])
	ee_pic_name  = json.loads(request.GET['ee_pic_name'])
	ee_remarkx   = json.loads(request.GET['ee_remarkx'])	

	_update  = models.remainder_details.objects.filter(user_id=ex_selectID).first()	
	_update.user_id 		= ex_selectID
	_update.user_name  		= ex_user_name	
	_update.remarks 		= ee_remarkx	
	_update.save()
	return HttpResponse(json.dumps('done'))


def view_pwd(request):
	user_list    = models.User_entry.objects.all().order_by('user_name')
	user_details = []
	for x in user_list:
		user_details.append({
			'user_id'   : x.id,
			'user_name' : x.user_name
			})

	view_pwd_details = models.view_password.objects.all()
	pwd_array = []
	for i in view_pwd_details:
		try:
			usr_name = i.user.user_name
		except:
			usr_name = ''
		pwd_array.append({
			'id' : i.id,
			'user_name' : usr_name,
			'view_pwd'  : i.view_password,
			'gen_date'  : i.generate_date,
			'mail_det'  : i.mail,
			'comments'  : i.comments,
			})
		

	context={
	'user_details' : user_details,
	'pwd_array'    : pwd_array,
	}

	return render_to_response("audit_display/view_pwd.html",context)


def update_pwd_details(request):
	ex_userID 			= json.loads(request.GET['ex_userID'])
	ex_user_name 		= json.loads(request.GET['ex_user_name'])
	ex_view_password 	= json.loads(request.GET['ex_view_password'])
	ex_comments 		= json.loads(request.GET['ex_comments'])
	ex_mail     		= json.loads(request.GET['ex_mail'])

	pwd_update 			= models.view_password.objects.filter(id=ex_userID).first()
	pwd_update.view_password = ex_view_password
	pwd_update.generate_date = datetime.now().date()
	pwd_update.comments      = ex_comments
	pwd_update.mail      	 = ex_mail
	pwd_update.save()
	return HttpResponse(json.dumps('done'))

def submit_pwd_details(request):
	e_user_name = json.loads(request.GET['e_userName'])
	e_view_pwd  = json.loads(request.GET['e_view_pwd'])
	e_remark    = json.loads(request.GET['e_remark'])
	e_mail      = json.loads(request.GET['e_mail'])
	user_det    = models.User_entry.objects.filter(user_name=e_user_name).first()
	chk_pwd	    = models.view_password.objects.filter(user_id=user_det).count()
	
	if chk_pwd==1:
	 	submit_pwd	= models.view_password.objects.filter(user_id=user_det).first()
	 	msg = 'no'
	else:
	 	submit_pwd	= models.view_password()
	 	msg = 'done'

	submit_pwd.user_id 		 = user_det.id
	submit_pwd.view_password = e_view_pwd
	submit_pwd.generate_date = datetime.now().date()
	submit_pwd.comments      = e_remark
	submit_pwd.mail          = e_mail 
	submit_pwd.save()	
	return HttpResponse(json.dumps(msg))

	
# def send_mail_for_recharge(request):	
	
# 	return HttpResponse(json.dumps(msg))

def contact_details(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		contact_details    	= models.emergency_contact.objects.all().order_by('user_name')
		info_array = []
		for x in contact_details:
			info_array.append({
				'id' 				: x.id,
				'user_name'			: x.user_name,
				'contact_mobile'    : x.contact_mobile,
				'contact_landline'  : x.contact_landline,
				'address'			: x.address,
				'remarks'			: x.remarks,
				'department'		: x.department
				})

		context={
		'info_array' : info_array
		}		
		return render_to_response("audit_display/contact_details.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/audit/user_login')
	else:
		return HttpResponseRedirect('/audit/user_login')


def submit_contact_entry(request):	
	user 				= request.GET['user']
	contact_mobile   	= request.GET['contact_mobile']
	contact_landline 	= request.GET['contact_landline']
	address 			= request.GET['address']	
	remark  			= request.GET['remark']
	department  		= request.GET['department']

	date_time 			= datetime.now().date()		
	user_already    	= models.emergency_contact.objects.filter(user_name=user).count()	
	if user_already==1:	
		userx    	= models.emergency_contact.objects.filter(user_name=user).first()		
		msg = 'already'
	else:
		userx 				  = models.emergency_contact()
		msg = 'done'
	
	userx.user_name 		= user
	userx.contact_mobile   	= contact_mobile
	userx.contact_landline 	= contact_landline
	userx.address 		  	= address	
	userx.remarks 		  	= remark
	userx.department 	 	= department
	userx.entry_date       	= date_time
	userx.save()
	return HttpResponse(json.dumps(msg))

def update_contact_entry(request):
	userID				= json.loads(request.GET['selectID'])
	user_name			= json.loads(request.GET['user'])	
	ex_contact_mobile	= json.loads(request.GET['ex_contact_mobile'])
	ex_contact_landline	= json.loads(request.GET['ex_contact_landline'])
	remark				= json.loads(request.GET['remark'])
	ex_address			= json.loads(request.GET['ex_address'])	
	ex_department		= json.loads(request.GET['ex_department'])
	
	update_user_entry 					= models.emergency_contact.objects.filter(id=userID).first()
	update_user_entry.user_name 	 	= user_name
	update_user_entry.contact_mobile  	= ex_contact_mobile
	update_user_entry.contact_landline 	= ex_contact_landline
	update_user_entry.address 		 	= ex_address
	update_user_entry.remarks 	     	= remark
	update_user_entry.department 		= ex_department	
	update_user_entry.save()
	return HttpResponse(json.dumps('done'))

def contact_details_download(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	
	file_name 						= "contact_details.xlsx"
	my_dir 	 	   					= os.path.dirname(__file__)
	location 	   					= (my_dir+'/'+file_name)
	filename 	   					= (my_dir+'/'+file_name)
	download_name  					= file_name
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response

@csrf_exempt
def export_contact_details(request):
	from shutil import copyfile
	import shutil

	fleet_details = json.loads(request.POST['user_details'])
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/contact_details.xlsx'
	dstroot       = select_dir+'/contact_details.xlsx'
	copyfile(srcfile, dstroot)
	my_dir 	      = os.path.dirname(__file__)
	location      = (my_dir+'/contact_details.xlsx')
	wb 	   	      = load_workbook(my_dir+'/contact_details.xlsx')
	ws     	      = wb.get_sheet_by_name("Sheet1")
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Arial', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='right'))
	align_style   = Style(font=Font(name='Arial', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='left'))	 	
	file_name     = 'contact_details.xlsx'
	#print '---------',fleet_details
	i = 3
	j = 1
	for ft in fleet_details:			
		user_name     	  = ft['Name']
		contact_mobile 	  = ft['Contact Number (Mobile)']		
		contact_landline  = ft['Contact Number (Landline)']	
		department 		  = ft['Department']
		address 		  = ft['Address']
		remarks 		  = ft['Remarks']			
		ws.cell('A'+str(i)).value = j
		ws.cell('B'+str(i)).value = user_name
		ws.cell('C'+str(i)).value = contact_mobile
		ws.cell('D'+str(i)).value = contact_landline
		ws.cell('E'+str(i)).value = department
		ws.cell('F'+str(i)).value = address
		ws.cell('G'+str(i)).value = remarks
		ws.cell('A'+str(i)).style = align_style
		ws.cell('B'+str(i)).style = align_style
		ws.cell('C'+str(i)).style = align_style
		ws.cell('D'+str(i)).style = align_style
		ws.cell('E'+str(i)).style = align_style
		ws.cell('F'+str(i)).style = align_style
		ws.cell('G'+str(i)).style = align_style		
		i+=1
		j+=1
	j+=1

	wb.save(my_dir+'/'+file_name)
	return HttpResponse(json.dumps('done'))
	

def system_information(request):

	return render_to_response("audit_display/system_information.html")

def get_alert_system(request):
	userID 			= json.loads(request.GET['userID'])
	gt_user_details = models.weekly_entry.objects.filter(id=userID)
	alert_list 		= []
	
	for x in gt_user_details:
		audited = x.entry_date.strftime('%d-%b-%Y')		
		try:
			user_det    = models.User_entry.objects.filter(id=x.user_id).first()
			assign_name = user_det.assigned_to
		except:
			assign_name = ''
		alert_list.append({
			'user_name' 	: x.user.user_name,
			'e_virus'		: x.up_antivirus,
			'e_cache'		: x.history_clean,
			'e_server'		: x.server_folder,
			'e_internet'	: x.internet_conn,
			'e_printer'		: x.printer_conn,
			'e_headphone'	: x.head_phone,
			'e_size'		: x.email_configure,
			'e_assigned_to' : assign_name,#x.assigned_to,
			'e_remarks'		: x.issue_any,
			'e_audited'		: str(audited),
			})

	return HttpResponse(json.dumps(alert_list))


def get_configuration_details(request):
	get_user_id    	= request.GET['user_id']
	weekly_details 	= models.User_entry.objects.filter(id=get_user_id)
	config_array   	= []
	ups_rep 		= None
	mouse_rep 		= None
	ssd_rep 		= None
	color_ups 		= ''
	color_mouse     = ''
	color_ssd 		= ''
	color_smps 		= ''
	color_pc 		= ''
	pc_rep 			= None
	color_kb 		= ''
	kb_rep 			= None
	monitor_rep     = None
	color_monitor 	= ''
	ram_rep 		= None
	color_ram 		= ''
	remarks 		= ''
	processor_rep   = None
	color_processor = ''

	for i in weekly_details:
		w = models.equipment_details.objects.filter(label=i.user_name)
		if i.other_items!=None and i.other_items!="":
			other_itemxs = i.other_items
		else:
			other_itemxs = ''

		if i.items=='Desktop':
			img   = '/static/img/pc.png'
		elif i.items=='Reception':
			img   = '/static/img/reception.jpg'			
		elif i.items=='Laptop':
			img   = '/static/img/laptop.png'			
		elif i.items=='DP Machine':
			img   = '/static/img/dp.png'
		elif i.items=='Server':
			img   = '/static/img/server.jpg'	
		elif i.items=='Camera':
			img   = '/static/img/camera.png' 
		elif i.items=='CCTV':
			img   = '/static/img/cctv.png' 				
		else:
			img   = '/static/img/blank.png'

		if i.other_items!=None and i.other_items!="#" and i.other_items!="":
			k_img = '/static/img/headphone.jpg'
		else:	
			k_img = '/static/img/blank.png'

		try:
			assigned_to = i.assigned_to
		except:
			assigned_to = ''

		try:
			dop = str(i.date_purchase).strftime('%d-%b-%Y')
		except:
			dop = ''


		for v in w:			
			# try:
			# 	if v.purchase_date!=None:
			# 		data_rep = str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
			# 	if v.repaired_date!=None:
			# 		data_rep = str(datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' R'
				
				
			# except:
			#  	data_rep = None
		

			try:
				if v.category=='UPS':
					if v.purchase_date==None or v.repaired_date!=None:
						ups_rep = str(datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' R'
						color_ups = 'yellow'					
					elif v.purchase_date!=None or v.repaired_date==None:
						ups_rep = str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
						color_ups = 'yellow'	
					else:
						ups_rep = None
						color_ups = ''	
			except:
				ups_rep = None
				color_ups = ''	


			try:
				if v.category=='Mouse':				
					if v.purchase_date==None and v.repaired_date!=None:
						mouse_rep = datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y')
						color_mouse = 'yellow'	
					elif v.purchase_date!=None and v.repaired_date==None:
						mouse_rep = str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
						color_mouse = 'yellow'	
					else:
						mouse_rep = None
						color_mouse = ''
			except:
				mouse_rep = None
				color_mouse = ''
			
			try:
				if v.category=='SSD':				
					if v.purchase_date==None and v.repaired_date!=None:
						ssd_rep 	= datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y')
						color_ssd 	= 'yellow'
					elif v.purchase_date!=None and v.repaired_date==None:
						ssd_rep 	= str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
						color_ssd 	= 'yellow'
					else:
						ssd_rep 	= None
						color_ssd 	= ''
			except:
				ssd_rep 	= None
				color_ssd 	= ''


			try:
				if v.category=='Keyboard':				
					if v.purchase_date==None and v.repaired_date!=None:
						kb_rep 	= datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y')
						color_kb 	= 'yellow'
					elif v.purchase_date!=None and v.repaired_date==None:
						kb_rep 	= str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
						color_kb 	= 'yellow'
					else:
						kb_rep 	= None
						color_kb 	= ''
			except:
				kb_rep 	= None
				color_kb 	= ''


			try:
				if v.category=='PC':				
					if v.purchase_date==None or v.repaired_date!=None:
						pc_rep 	= datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y')
						color_pc 	= 'yellow'
					elif v.purchase_date!=None or v.repaired_date==None:
						pc_rep 	= str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
						color_pc 	= 'yellow'
					else:
						pc_rep 	= None
						color_pc 	= ''
			except:
				pc_rep 	= None
				color_pc 	= ''


			try:
				if v.category=='Monitor':				
					if v.purchase_date==None or v.repaired_date!=None:
						monitor_rep 	= datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y')
						color_monitor 	= 'yellow'
					elif v.purchase_date!=None or v.repaired_date==None:
						monitor_rep 	= str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
						color_monitor 	= 'yellow'
					else:
						monitor_rep 	= None
						color_monitor 	= ''
			except:
				monitor_rep 	= None
				color_monitor 	= ''


			try:
				if v.category=='RAM':
					if v.purchase_date==None or v.repaired_date!=None:
						ram_rep = str(datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' R'
						color_ram = 'yellow'					
					elif v.purchase_date!=None or v.repaired_date==None:
						ram_rep = str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
						color_ram = 'yellow'	
					else:
						ram_rep = None
						color_ram = ''	
			except:
				ram_rep = None
				color_ram = ''	


			try:
				if v.category=='Processor':
					if v.purchase_date==None or v.repaired_date!=None:
						processor_rep = str(datetime.strptime(str(v.repaired_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' R'
						color_processor = 'yellow'					
					elif v.purchase_date!=None or v.repaired_date==None:
						processor_rep = str(datetime.strptime(str(v.purchase_date),"%Y-%m-%d").strftime('%d-%b-%Y'))+' B'
						color_processor = 'yellow'	
					else:
						processor_rep = None
						color_processor = ''	
			except:
				processor_rep = None
				color_processor = ''	

			try:
				remarks = v.remarks
			except:
				remarks = ''

		


		
		config_array.append({
			'userID'             : i.id,
			'user_name'          : i.user_name,			
			'user_id'            : i.id,
			'info_hdd'		     : i.hdd,
			'info_ram'		     : i.ram,
			'info_processor'     : i.processor,
			'info_operating_sys' : i.operating_sys,
			'info_remarks'  	 : i.remark,
			'info_keyboard' 	 : i.keyboard,
			'info_mouse'         : i.mouse,
			'info_smps' 		 : i.smps,
			'info_monitor'       : i.monitor,
			'img' 				 : '', #img
			'k_img' 			 : k_img,
			'serial_no'			 : other_itemxs,
			'assgined_for'		 : assigned_to,
			'location'			 : i.location,
			'date_of_purchase'   : dop,
			'ups_rep'			 : ups_rep,
			'mouse_rep' 		 : mouse_rep,
			'ssd_rep' 			 : ssd_rep,
			'color_ups'			 : color_ups,
			'color_mouse' 		 : color_mouse,
			'pc_rep' 			 : pc_rep,
			'color_pc' 		 	 : color_pc,
			'kb_rep' 			 : kb_rep,
			'color_kb' 			 : color_kb,
			'monitor_rep' 		 : monitor_rep,
			'color_monitor' 	 : color_monitor,
			'ram_rep' 			 : ram_rep,
			'color_ram' 		 : color_ram,
			'remarks' 			 : remarks,
			'color_ssd'			 : color_ssd,
			'processor_rep'		 : processor_rep,
			'color_processor'	 : color_processor,
		})

	context={
		'config_array'  : config_array
	}

	return HttpResponse(json.dumps(context))